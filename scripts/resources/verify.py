#!/usr/bin/env python3
"""Independent verification of every WEI worksheet formula.

For each XLSX we inject a sample scenario into the input cells, recompute the
workbook with the `formulas` engine (NOT exceljs, which only stores formulas),
and assert the computed outputs equal values worked out by hand here. This is a
real end to end check that the shipped formulas are correct."""
import os, glob, tempfile, warnings, logging
import openpyxl, formulas

logging.disable(logging.CRITICAL); warnings.filterwarnings("ignore")
DIR = os.path.join(os.path.dirname(__file__), "..", "..", "public", "resources", "files")
FAILS = []

def row_of(ws, text, col=1):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, col).value).strip() == text:
            return r
    raise KeyError(text)

def calc(path, inputs):
    wb = openpyxl.load_workbook(path); ws = wb.active
    for addr, val in inputs.items():
        ws[addr] = val
    tmp = tempfile.mktemp(suffix=".xlsx"); wb.save(tmp)
    xl = formulas.ExcelModel().loads(tmp).finish()
    sol = xl.calculate()
    fn = os.path.basename(tmp); sheet = ws.title.upper()
    def get(addr):
        v = sol["'[%s]%s'!%s" % (fn, sheet, addr)].value
        try: return v[0][0]
        except Exception: return v
    os.remove(tmp)
    return get

def check(name, got, exp, tol=0.02):
    ok = (isinstance(got, (int, float)) and abs(got - exp) <= tol) or got == exp
    print(("  PASS " if ok else "  FAIL ") + "%-46s got=%s exp=%s" % (name, got, exp))
    if not ok: FAILS.append((name, got, exp))

def f(n): return os.path.join(DIR, n)

# 1. BUDGET -----------------------------------------------------------------
print("\n# Monthly budget worksheet")
g = calc(f("wei-monthly-budget-worksheet.xlsx"),
         {"B7": 4000, "B18": 1200, "C18": 1250, "B19": 150, "C19": 140})
check("needs target (50% of 4000)", g("C11"), 2000)
check("wants target (30%)", g("C12"), 1200)
check("savings target (20%)", g("C13"), 800)
check("rent difference 1200-1250", g("D18"), -50)
check("utilities difference 150-140", g("D19"), 10)
check("needs subtotal planned", g("B25"), 1350)
check("needs subtotal actual", g("C25"), 1390)
wb = openpyxl.load_workbook(f("wei-monthly-budget-worksheet.xlsx")); ws = wb.active
tot_row = row_of(ws, "Total planned")
lb_row = row_of(ws, "Left to budget (income minus planned)")
check("total planned", g("B%d" % tot_row), 1350)
check("left to budget 4000-1350", g("B%d" % lb_row), 2650)

# 2. SAVINGS ----------------------------------------------------------------
print("\n# Savings goal tracker")
wb = openpyxl.load_workbook(f("wei-savings-goal-tracker.xlsx")); ws = wb.active
tgt = row_of(ws, "Target amount"); start = row_of(ws, "Amount you already have saved")
mo = row_of(ws, "Amount you can save each month")
still = row_of(ws, "Amount still needed"); months = row_of(ws, "Months to reach your goal at this pace")
log_hdr = row_of(ws, "Date"); l1 = log_hdr + 1
g = calc(f("wei-savings-goal-tracker.xlsx"),
         {"B%d" % tgt: 1000, "B%d" % start: 200, "B%d" % mo: 100,
          "B%d" % l1: 50, "B%d" % (l1 + 1): 75})
check("still needed 1000-200", g("B%d" % still), 800)
check("months to goal ceil(800/100)", g("B%d" % months), 8)
check("balance after first contribution", g("C%d" % l1), 250)
check("balance after second", g("C%d" % (l1 + 1)), 325)
check("percent of goal row1 250/1000", g("D%d" % l1), 0.25)

# 3. DEBT -------------------------------------------------------------------
print("\n# Debt payoff planner")
g = calc(f("wei-debt-payoff-planner.xlsx"),
         {"B8": 1000, "C8": 20, "D8": 50, "B9": 4000, "C9": 6, "D9": 80,
          "B10": 500, "C10": 25, "D10": 25, "B19": 100})
check("monthly interest 1000*20%/12", g("E8"), 1000 * 0.20 / 12)
check("monthly interest 4000*6%/12", g("E9"), 20)
check("avalanche order (25% APR -> 1)", g("F10"), 1)
check("avalanche order (20% APR -> 2)", g("F8"), 2)
check("snowball order (500 bal -> 1)", g("G10"), 1)
check("snowball order (1000 bal -> 2)", g("G8"), 2)
check("total balance", g("B16"), 5500)
check("total minimum payments", g("D16"), 155)
check("total toward debt 155+100", g("B20"), 255)

# 4. (paycheck is PDF only, no formulas)

# 5. BANK -------------------------------------------------------------------
print("\n# Bank account comparison")
g = calc(f("wei-bank-account-comparison.xlsx"),
         {"B10": 12, "B12": 0.5, "B18": 2000, "C10": 0, "C12": 4.0})
check("yearly fees 12*12", g("B21"), 144)
check("yearly interest 2000*0.5%", g("B22"), 10)
check("net first year 10-144", g("B23"), -134)
check("acct2 yearly interest 2000*4%", g("C22"), 80)

# 6. EMERGENCY --------------------------------------------------------------
print("\n# Emergency fund planner")
g = calc(f("wei-emergency-fund-planner.xlsx"),
         {"B8": 1000, "B9": 200, "B10": 400, "B21": 3, "B25": 1200, "B26": 300})
check("essentials sum", g("B16"), 1600)
check("three months", g("B19"), 4800)
check("fund target 1600*3", g("B22"), 4800)
check("still needed 4800-1200", g("B27"), 3600)
check("months to target ceil(3600/300)", g("B28"), 12)
check("progress 1200/4800", g("B29"), 0.25)

# 7. COLLEGE ----------------------------------------------------------------
print("\n# College cost / net price")
g = calc(f("wei-college-cost-net-price-worksheet.xlsx"),
         {"B8": 10000, "B9": 12000, "B10": 1200, "B11": 1500, "B12": 2000,
          "B17": 8000, "B18": 5000, "B25": 3000, "B26": 2000, "B27": 2500, "B28": 5000})
check("cost of attendance sum", g("B13"), 26700)
check("total gift aid", g("B20"), 13000)
check("net price 26700-13000", g("B21"), 13700)
check("total resources", g("B30"), 12500)
check("remaining gap 13700-12500", g("B31"), 1200)
check("four-year 13700*4", g("B35"), 54800)

# 8. SPENDING ---------------------------------------------------------------
print("\n# Spending tracker")
wb = openpyxl.load_workbook(f("wei-spending-tracker.xlsx")); ws = wb.active
hdr = row_of(ws, "Date"); s1 = hdr + 1
foodtot = row_of(ws, "Food")
g = calc(f("wei-spending-tracker.xlsx"),
         {"C%d" % s1: "Food", "E%d" % s1: 20, "D%d" % s1: "Need",
          "C%d" % (s1 + 1): "Food", "E%d" % (s1 + 1): 10, "D%d" % (s1 + 1): "Want",
          "C%d" % (s1 + 2): "Housing", "E%d" % (s1 + 2): 600, "D%d" % (s1 + 2): "Need",
          "C%d" % (s1 + 3): "Fun", "E%d" % (s1 + 3): 15, "D%d" % (s1 + 3): "Want"})
check("food total via SUMIF", g("B%d" % foodtot), 30)
check("housing total via SUMIF", g("B%d" % row_of(ws, "Housing")), 600)
need = row_of(ws, "Total needs"); want = row_of(ws, "Total wants")
check("needs total 600+20", g("B%d" % need), 620)
check("wants total 10+15", g("B%d" % want), 25)

print("\n" + ("ALL PASS" if not FAILS else "FAILURES: %d" % len(FAILS)))
raise SystemExit(1 if FAILS else 0)
